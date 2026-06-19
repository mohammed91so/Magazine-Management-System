"""
Unit tests for reporting service exports.
"""

from openpyxl import load_workbook

from services.inventory_service import create_product
from services.reporting_service import ReportingService
from services.sales_service import sell_product


class TestReportingService:
    """Test reporting service exports."""

    def test_export_combined_report_includes_dashboard_metrics(self, temp_db, tmp_path):
        product_id = create_product("Report Product", 10.0, 15.0, 50, "2030-12-31")
        sell_product(product_id, 5)
        reporting_service = ReportingService(output_dir=tmp_path)

        output_path = reporting_service.export_combined_report("combined_report.xlsx")

        assert output_path.exists()

        workbook = load_workbook(output_path)
        assert workbook.sheetnames == ["Sales", "Inventory", "Dashboard"]

        dashboard_sheet = workbook["Dashboard"]
        dashboard_metrics = {
            row[0]: row[1]
            for row in dashboard_sheet.iter_rows(min_row=2, values_only=True)
        }

        assert dashboard_metrics["total_earnings"] == 75
        assert dashboard_metrics["total_profit"] == 25
        assert dashboard_metrics["total_products"] == 1
